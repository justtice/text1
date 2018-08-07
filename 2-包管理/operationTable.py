#!/usr/bin/python
# -*- coding: utf-8 -*-
#coding=utf-8
import openpyxl
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import Border, Side, Font, colors

class Excel():

    def pathName(self,path):

        # workbook_name = '0806黄爱枝(1).xlsx'
        # path = "/Users/liulunyang/Desktop/0806黄爱枝.xlsx"
        wb = openpyxl.load_workbook(path)

        # wb = Workbook()
        ws = wb.active
        # ws["A1"] = 0
        # ws.remove_named_range()
        # help(ws)
        # ws.merge_cells('A2:B2')

        # 需要先取消合并
        max_row = ws.max_row
        print(max_row)
        ws.unmerge_cells(start_row=max_row, start_column=1, end_row=max_row, end_column=3)
        ws.unmerge_cells(start_row=max_row, start_column=5, end_row=max_row, end_column=8)
        ws.unmerge_cells(start_row=max_row, start_column=10, end_row=max_row, end_column=14)
        ws.unmerge_cells(start_row=max_row - 1, start_column=1, end_row=max_row - 1, end_column=3)
        ws.unmerge_cells(start_row=max_row - 1, start_column=5, end_row=max_row - 1, end_column=8)
        ws.unmerge_cells(start_row=max_row - 1, start_column=10, end_row=max_row - 1, end_column=14)
        ws.unmerge_cells('A1:O1')

        # 删除列
        for i in range(4):
            ws.delete_cols(1)

        ws.delete_cols(2)
        ws.delete_cols(8)
        ws.delete_cols(8)
        # 删除第一行
        ws.delete_rows(1)
        ws.delete_rows(max_row - 2)

        # 改变列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 3.8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 11

        '''
        这方法对一行不起作用
        ws.column_dimensions['A'].font = Font(bold=True,size=15)
        ws['A1'].font = Font(size = 20)
        '''

        ws.cell(row=max_row - 2, column=1).value = "总计:"
        # print(ws['D2'].value,len(ws['D2'].value))
        ws.row_dimensions[max_row - 2].height = 15
        for i in range(max_row - 3):
            if i == max_row - 4:
                break
            ws.cell(row=i + 1, column=1).font = Font(size=12, bold=True)
            ws.cell(row=i + 1, column=1).alignment = styles.Alignment(vertical='center', horizontal='center')
            a = ws.cell(row=i + 2, column=1).value
            print(a, len(a))
            d = ws.cell(row=i + 2, column=4).value
            print(d)
            if d == 0:
                d = 1

            if len(a) < 14:
                if d > 1:
                    ws.row_dimensions[i + 2].height = 13 * d
                else:
                    ws.row_dimensions[i + 2].height = 15 * d
            else:
                if d > 2:
                    ws.row_dimensions[i + 2].height = 13 * d
                else:
                    ws.row_dimensions[i + 2].height = 15 * 2
        ## 有问题这里, 改变cell的背景颜色
        # ws['A1'].styles=styles.PatternFill(fill_type='solid',fgColor="0d5330")
        # ws.row_dimensions[1].fill = styles.PatternFill(bgColor=colors.BLACK)
        '''
        元格样式
        font(字体类)：字号、字体颜色、下划线等
        fill(填充类)：颜色等
        border(边框类)：设置单元格边框
        alignment(位置类)：对齐方式
        number_format(格式类)：数据格式
        protection(保护类)：写保护
        '''

        fill = styles.PatternFill(fill_type=None,
                                  bgColor=colors.WHITE)

        # dashDot','dashDotDot', 'dashed','dotted',
        #                             'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
        #                             'mediumDashed', 'slantDashDot', 'thick', 'thin'
        # 修改边框
        border = Border(left=Side(style='hair', color=colors.BLACK), right=Side(style='hair', color=colors.BLACK),
                        top=Side(style="hair", color=colors.BLACK), bottom=Side(style='hair', color=colors.BLACK))
        ws['D1'] = '数量'

        for r in range(max_row - 2):
            for i in range(7):
                # ws.cell(row= r + 1, column =1+i).fill = styles.PatternFill(bgColor=colors.BLACK)
                ws.cell(row=r + 1, column=1 + i).fill = fill
                ws.cell(row=r + 1, column=1 + i).border = border
                ws.cell(row=r + 1, column=1).alignment = styles.Alignment(wrap_text=True)

        # 改变第一行的居中对齐
        for i in range(7):
            ws.cell(row=1, column=i + 1).alignment = styles.Alignment(vertical='center', horizontal='center')
            ws.cell(row=1, column=1 + i).font = Font(size=12, bold=True, b=1)

        wb.save(path)
        # 自适应行高   没有实现

        # help(openpyxl.styles)


        '''
        编辑打印模式
        from openpyxl.workbook import Workbook

        wb = Workbook()
        ws = wb.active

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1

        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        '''

        #2、pycharm使多行代码同时左移
        #            鼠标选中多行代码后，同时按住shift+Tab键，一次左移四个字符